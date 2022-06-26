from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Read workbook and select sheet
wb = load_workbook('/Users/raulgiron/Desktop/AutomateWithPython/pivot_table_barchart.xlsx')
sheet = wb['Report']

# Active rows and columns
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# sheet['B8'] = '=SUM(B6:B7)'
# sheet['B8'].style = 'Comma'
for i in range(min_column+1, max_column+1):
    letter = (get_column_letter(i))
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Comma'
    # print(i)
    # print(f'=SUM({letter}{min_row+1}:{letter}{max_row})')

wb.save('/Users/raulgiron/Desktop/AutomateWithPython/pivot_table_barchart.xlsx')
